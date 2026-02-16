# -*- coding: utf-8 -*-
import io
import json
import re
import tkinter as tk
import unicodedata
from dataclasses import dataclass, asdict, field
from pathlib import Path
from tkinter import filedialog, messagebox

import msoffcrypto
from openpyxl import load_workbook


# =========================
#  設定（環境ファイルに保存）
# =========================
@dataclass
class JudgerConfig:
    # Step1: pXXX評定.xlsx の検索ディレクトリ
    target_dir: str = ""

    # 現実の命名揺れに対応（例: p25101280 評定.xlsx）
    filename_regex: str = r"^p\d+.*評定.*\.xlsx$"

    # 互換のため残す（連番指定）
    chapter_start: int = 23
    chapter_count: int = 2

    # ★非連番対応：ここが空でなければこちらを優先
    chapter_numbers: list[int] = field(default_factory=list)

    # 合否閾値
    pass_score_threshold: float = 4.0
    permission_threshold: float = 1.0

    # 不合格条件（どれかの章が <=3）
    fail_if_any_score_leq: float = 3.0

    # ★セルに書き込むのは G/F/5 のみ（固定運用）
    label_pass: str = "G"
    label_fail: str = "F"
    label_skip: str = "5"

    # Step1: 判定列
    output_header: str = "判定"
    output_suffix: str = "_判定付き"

    # Step2: seiseki側（パスワード付き）
    seiseki_path: str = ""  # 空なら target_dir/seiseki.xlsx を使う
    seiseki_password: str = ""  # GUI入力

    seiseki_student_header: str = "学生番号"
    seiseki_eval_header: str = "評価入力欄（G:合格 F:不合格 5:評価せず）"

    judged_elms_header: str = "ELMS ID"
    judged_result_header: str = "判定"  # output_header と一致推奨

    # Step2 出力
    seiseki_output_suffix: str = "_判定反映"

    # 設定ファイル
    config_filename: str = "judge_config.json"

    # 未一致表示をログに出す最大件数
    max_log_missing: int = 30


# =========================
#  正規化・ヘッダーマッチ
# =========================
def normalize_filename(name: str) -> str:
    return unicodedata.normalize("NFKC", name)


def normalize_text(s: str) -> str:
    if s is None:
        return ""
    s = unicodedata.normalize("NFKC", str(s))
    s = s.replace(" ", "").replace("\u3000", "")
    s = s.replace("(", "").replace(")", "").replace("（", "").replace("）", "")
    return s.lower()


def extract_digits(s) -> str:
    """例: s20243025a -> 20243025 / 20243025 -> 20243025"""
    if s is None:
        return ""
    t = unicodedata.normalize("NFKC", str(s))
    digits = re.findall(r"\d+", t)
    return "".join(digits)


def find_col_by_exact_header(headers: list[str], header_name: str) -> int | None:
    """完全一致（NFKC+strip）でヘッダー列を探す"""
    target = unicodedata.normalize("NFKC", header_name).strip()
    for idx, h in enumerate(headers, start=1):
        if unicodedata.normalize("NFKC", str(h) if h is not None else "").strip() == target:
            return idx
    return None


# =========================
#  章番号指定のパース
# =========================
def parse_chapter_numbers(spec: str) -> list[int]:
    """
    章番号指定をパースする。
    例:
      "23,25,30" -> [23,25,30]
      "23-25, 30" -> [23,24,25,30]
      "23 25 30" -> [23,25,30]
      "23?25" -> [23,24,25]
    """
    if spec is None:
        return []
    s = unicodedata.normalize("NFKC", str(spec)).strip()
    if s == "":
        return []

    s = s.replace("、", ",").replace(";", ",")
    parts = re.split(r"[,\s]+", s)

    nums: list[int] = []
    for p in parts:
        if p == "":
            continue
        p2 = p.replace("?", "-").replace("?", "-").replace("?", "-")
        if "-" in p2:
            a_b = p2.split("-")
            if len(a_b) != 2:
                raise ValueError(f"章番号リストの範囲指定が不正: '{p}'")
            a = int(a_b[0])
            b = int(a_b[1])
            if a <= b:
                nums.extend(list(range(a, b + 1)))
            else:
                nums.extend(list(range(a, b - 1, -1)))
        else:
            nums.append(int(p2))

    # 重複除去（順序維持）
    seen = set()
    out = []
    for n in nums:
        if n not in seen:
            out.append(n)
            seen.add(n)
    return out


def get_chapters(cfg: JudgerConfig) -> list[int]:
    if cfg.chapter_numbers:
        return list(cfg.chapter_numbers)
    return [cfg.chapter_start + i for i in range(cfg.chapter_count)]


def safe_cfg_from_dict(data: dict) -> JudgerConfig:
    """
    古いjsonに余計なキーが入っていても落ちないように、
    dataclassのフィールドだけ拾って生成する。
    """
    allowed = {f.name for f in JudgerConfig.__dataclass_fields__.values()}
    filtered = {k: v for k, v in data.items() if k in allowed}
    return JudgerConfig(**filtered)


# =========================
#  欠損判定・数値化
# =========================
def is_missing(v) -> bool:
    if v is None:
        return True
    if isinstance(v, str):
        if v.strip() == "" or v.strip() == "-":
            return True
    return False


def to_float(v):
    if is_missing(v):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, str):
        try:
            return float(v.strip())
        except ValueError:
            return None
    return None


# =========================
#  判定ロジック（Step1）
#  ※英語列が存在しない／日本語列が存在しない、両方に対応
# =========================
def judge_row(
    *,
    row_values: list,
    jp_ch_cols: list[int] | None,
    jp_perm_col: int | None,
    en_ch_cols: list[int] | None,
    en_perm_col: int | None,
    cfg: JudgerConfig
) -> str:
    # セットが「存在する」かどうか（列が揃っているか）
    jp_exists = (jp_ch_cols is not None) and (jp_perm_col is not None)
    en_exists = (en_ch_cols is not None) and (en_perm_col is not None)

    if not jp_exists and not en_exists:
        raise ValueError("日本語版・英語版のいずれの評価列も見つかりません（ファイルのヘッダーを確認してください）")

    # その行に入力があるか
    jp_any_present = False
    en_any_present = False
    if jp_exists:
        jp_any_present = any(not is_missing(row_values[c]) for c in (jp_ch_cols + [jp_perm_col]))
    if en_exists:
        en_any_present = any(not is_missing(row_values[c]) for c in (en_ch_cols + [en_perm_col]))

    # どちらにも入力がない（=全部-や空）なら評価せず
    if not jp_any_present and not en_any_present:
        return cfg.label_skip

    # 使う言語セットを決める
    if jp_exists and jp_any_present and (not en_any_present):
        use_jp = True
    elif en_exists and en_any_present and (not jp_any_present):
        use_jp = False
    else:
        # 両方に何か入ってる or 片方が存在しないが片方が present
        if not en_exists:
            use_jp = True
        elif not jp_exists:
            use_jp = False
        else:
            # 両方存在して両方present：欠損が少ない方（同数なら日本語）
            jp_missing = sum(1 for c in (jp_ch_cols + [jp_perm_col]) if is_missing(row_values[c]))
            en_missing = sum(1 for c in (en_ch_cols + [en_perm_col]) if is_missing(row_values[c]))
            use_jp = (jp_missing <= en_missing)

    if use_jp:
        ch_cols = jp_ch_cols
        perm_col = jp_perm_col
    else:
        ch_cols = en_ch_cols
        perm_col = en_perm_col

    if ch_cols is None or perm_col is None:
        return cfg.label_skip

    # 評価せず：必須のどれか欠損
    required_cols = ch_cols + [perm_col]
    if any(is_missing(row_values[c]) for c in required_cols):
        return cfg.label_skip

    # 数値化
    ch_scores = []
    for c in ch_cols:
        f = to_float(row_values[c])
        if f is None:
            return cfg.label_skip
        ch_scores.append(f)

    perm = to_float(row_values[perm_col])
    if perm is None:
        return cfg.label_skip

    # 合格
    pass_ok = all(s >= cfg.pass_score_threshold for s in ch_scores) and (perm >= cfg.permission_threshold)
    if pass_ok:
        return cfg.label_pass

    # 不合格：章のどれかが <=3
    fail_low = any(s <= cfg.fail_if_any_score_leq for s in ch_scores)
    if fail_low:
        return cfg.label_fail

    return cfg.label_fail


# =========================
#  ヘッダー解析（章番号抽出で列を確定）
# =========================
def build_column_maps(headers: list[str]):
    jp_ch_map = {}
    en_ch_map = {}
    jp_perm_col = None
    en_perm_col = None

    re_jp_ch = re.compile(r"第(\d+)章")
    re_en_ch = re.compile(r"chapter(\d+)")

    for idx, h in enumerate(headers, start=1):
        hn = normalize_text(h)

        # permission列（英）
        if "小テスト" in hn and "permission" in hn and "supervisor" in hn and "実データ" in hn:
            en_perm_col = idx

        # permission列（日）
        if "小テスト" in hn and "指導教員" in hn and "許可" in hn and "実データ" in hn:
            jp_perm_col = idx

        # 日本語章列
        if "小テスト" in hn and "確認テスト" in hn and "実データ" in hn:
            m = re_jp_ch.search(hn)
            if m:
                n = int(m.group(1))
                jp_ch_map[n] = idx

        # 英語章列
        if "小テスト" in hn and "confirmation" in hn and "実データ" in hn:
            m = re_en_ch.search(hn)
            if m:
                n = int(m.group(1))
                en_ch_map[n] = idx

    return jp_ch_map, en_ch_map, jp_perm_col, en_perm_col


def build_language_cols(chapters: list[int], ch_map: dict[int, int], perm_col: int | None):
    """
    chaptersに対応する列が全て存在し、perm_colもある場合のみ (cols, perm) を返す。
    どれか欠けるなら (None, None) を返す。
    """
    if perm_col is None:
        return None, None
    cols = []
    for n in chapters:
        if n not in ch_map:
            return None, None
        cols.append(ch_map[n])
    return cols, perm_col


# =========================
#  Step1: pXXX評定.xlsx -> 判定列付与
# =========================
def process_one_pfile(xlsx_path: Path, cfg: JudgerConfig) -> Path:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    chapters = get_chapters(cfg)

    jp_ch_map, en_ch_map, jp_perm_col, en_perm_col = build_column_maps(headers)

    # ★英語列が無い場合／日本語列が無い場合に対応
    jp_ch_cols, jp_perm_col2 = build_language_cols(chapters, jp_ch_map, jp_perm_col)
    en_ch_cols, en_perm_col2 = build_language_cols(chapters, en_ch_map, en_perm_col)

    if (jp_ch_cols is None or jp_perm_col2 is None) and (en_ch_cols is None or en_perm_col2 is None):
        raise ValueError(
            f"[{xlsx_path.name}] 日本語版・英語版いずれの評価列も揃いませんでした。"
            f" 章={chapters} / ヘッダー命名を確認してください。"
        )

    # 先頭列を挿入して結果を書き込む
    ws.insert_cols(1)
    ws.cell(row=1, column=1).value = cfg.output_header

    # 列シフト
    shift = 1
    if jp_ch_cols is not None and jp_perm_col2 is not None:
        jp_ch_cols = [c + shift for c in jp_ch_cols]
        jp_perm_col2 = jp_perm_col2 + shift
    else:
        jp_ch_cols = None
        jp_perm_col2 = None

    if en_ch_cols is not None and en_perm_col2 is not None:
        en_ch_cols = [c + shift for c in en_ch_cols]
        en_perm_col2 = en_perm_col2 + shift
    else:
        en_ch_cols = None
        en_perm_col2 = None

    for r in range(2, ws.max_row + 1):
        # row_values は「列番号で直接参照できるように」1-basedに揃える
        row_vals = [None] + [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        result = judge_row(
            row_values=row_vals,
            jp_ch_cols=jp_ch_cols,
            jp_perm_col=jp_perm_col2,
            en_ch_cols=en_ch_cols,
            en_perm_col=en_perm_col2,
            cfg=cfg
        )
        ws.cell(row=r, column=1).value = result

    out_path = xlsx_path.with_name(xlsx_path.stem + cfg.output_suffix + xlsx_path.suffix)
    wb.save(out_path)
    return out_path


def find_pfiles(cfg: JudgerConfig) -> list[Path]:
    target = Path(cfg.target_dir)
    pat = re.compile(cfg.filename_regex, flags=re.IGNORECASE)
    files = []
    for p in sorted(target.iterdir()):
        if not p.is_file():
            continue
        if p.name.startswith("~$"):
            continue
        if p.suffix.lower() != ".xlsx":
            continue
        if pat.match(normalize_filename(p.name)):
            files.append(p)
    return files


def find_judged_files(cfg: JudgerConfig) -> list[Path]:
    """Step2用：判定付きファイル（*_判定付き.xlsx）を探す"""
    target = Path(cfg.target_dir)
    files = []
    for p in sorted(target.iterdir()):
        if not p.is_file():
            continue
        if p.name.startswith("~$"):
            continue
        if p.suffix.lower() != ".xlsx":
            continue
        if p.stem.endswith(cfg.output_suffix):
            files.append(p)
    return files


# =========================
#  Step2: 判定付き -> seiseki.xlsx へ反映（未一致は警告して無視）
# =========================
def decrypt_xlsx_to_bytesio(path: Path, password: str) -> io.BytesIO:
    decrypted = io.BytesIO()
    with open(path, "rb") as f:
        office = msoffcrypto.OfficeFile(f)
        office.load_key(password=password)
        office.decrypt(decrypted)
    decrypted.seek(0)
    return decrypted


def build_id_to_result_map(judged_path: Path, cfg: JudgerConfig) -> tuple[dict[str, str], list[str]]:
    logs = []
    wb = load_workbook(judged_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_elms = find_col_by_exact_header(headers, cfg.judged_elms_header)
    col_result = find_col_by_exact_header(headers, cfg.judged_result_header)

    if col_elms is None:
        raise ValueError(f"[{judged_path.name}] ヘッダー '{cfg.judged_elms_header}' が見つかりません")
    if col_result is None:
        raise ValueError(f"[{judged_path.name}] ヘッダー '{cfg.judged_result_header}' が見つかりません")

    id2res: dict[str, str] = {}
    dup = 0
    empty = 0

    for r in range(2, ws.max_row + 1):
        elms = ws.cell(row=r, column=col_elms).value
        res = ws.cell(row=r, column=col_result).value

        sid = extract_digits(elms)
        if sid == "":
            empty += 1
            continue
        if sid in id2res:
            dup += 1
            continue

        res_s = "" if res is None else str(res).strip()

        # ★判定値はG/F/5のみ
        if res_s not in (cfg.label_pass, cfg.label_fail, cfg.label_skip):
            raise ValueError(
                f"[{judged_path.name}] 判定値が想定外です: ID={elms} -> '{res_s}' "
                f"(allowed: {cfg.label_pass},{cfg.label_fail},{cfg.label_skip})"
            )

        id2res[sid] = res_s

    logs.append(f"[{judged_path.name}] mapped IDs: {len(id2res)}, empty_elms: {empty}, duplicates_skipped: {dup}")
    return id2res, logs


def apply_results_to_seiseki(
    seiseki_path: Path,
    seiseki_password: str,
    id2res: dict[str, str],
    cfg: JudgerConfig
) -> tuple[Path, list[str], list[str]]:
    """
    返り値:
      out_path, logs, missing_ids
    missing_ids は「判定付きにあるが seiseki に存在しない（数字一致できない）」学生番号のリスト。
    """
    logs = []

    decrypted = decrypt_xlsx_to_bytesio(seiseki_path, seiseki_password)
    wb = load_workbook(decrypted, data_only=False)  # 書き込みするので data_only=False
    ws = wb[wb.sheetnames[0]]

    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_stu = find_col_by_exact_header(headers, cfg.seiseki_student_header)
    col_eval = find_col_by_exact_header(headers, cfg.seiseki_eval_header)

    if col_stu is None:
        raise ValueError(f"[{seiseki_path.name}] ヘッダー '{cfg.seiseki_student_header}' が見つかりません")
    if col_eval is None:
        raise ValueError(f"[{seiseki_path.name}] ヘッダー '{cfg.seiseki_eval_header}' が見つかりません")

    # seisekiの学生番号 -> 行番号
    stu2row: dict[str, int] = {}
    dup = 0
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=col_stu).value
        sid = extract_digits(v)
        if sid == "":
            continue
        if sid in stu2row:
            dup += 1
            continue
        stu2row[sid] = r

    missing_ids = [sid for sid in id2res.keys() if sid not in stu2row]

    logs.append(f"[{seiseki_path.name}] seiseki rows mapped: {len(stu2row)}, duplicates_skipped: {dup}")
    logs.append(f"Judged IDs: {len(id2res)}, missing_in_seiseki: {len(missing_ids)}")

    # ★未一致は「警告して無視」（停止しない）
    if missing_ids:
        show = missing_ids[: cfg.max_log_missing]
        logs.append("WARNING: seiseki側に存在しない学生番号（数字一致できず）がありました。該当学生は無視して続行します。")
        logs.append(f"Missing IDs (showing up to {cfg.max_log_missing}): {', '.join(show)}")

    # 一致したものだけ反映
    updated = 0
    skipped = 0
    for sid, res in id2res.items():
        if sid not in stu2row:
            skipped += 1
            continue
        r = stu2row[sid]
        ws.cell(row=r, column=col_eval).value = res
        updated += 1

    logs.append(f"Applied: {updated}, Skipped(no match): {skipped}")

    out_path = seiseki_path.with_name(seiseki_path.stem + cfg.seiseki_output_suffix + seiseki_path.suffix)

    # openpyxlでは再暗号化できない（ここは仕様）
    wb.save(out_path)
    logs.append(f"Saved (unencrypted): {out_path.name}")
    return out_path, logs, missing_ids


# =========================
#  GUI
# =========================
class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("評定ファイル 判定＆seiseki反映ツール（プロトタイプ）")
        self.geometry("980x740")
        self.cfg = JudgerConfig()
        self._build_ui()

    def _build_ui(self):
        frm = tk.Frame(self)
        frm.pack(fill="both", expand=True, padx=10, pady=10)

        tk.Label(frm, text="対象ディレクトリ（pXXX評定.xlsx と seiseki.xlsx）").grid(row=0, column=0, sticky="w")
        self.dir_var = tk.StringVar(value=self.cfg.target_dir)
        tk.Entry(frm, textvariable=self.dir_var, width=95).grid(row=0, column=1, sticky="we", padx=5)
        tk.Button(frm, text="選択", command=self.choose_dir).grid(row=0, column=2, padx=5)

        # 章番号リスト（優先）
        tk.Label(frm, text="章番号リスト（優先・例: 23,25 / 23-25,30 / 空なら下の連番）").grid(row=1, column=0, sticky="w")
        self.ch_list_var = tk.StringVar(value=",".join(map(str, self.cfg.chapter_numbers)))
        tk.Entry(frm, textvariable=self.ch_list_var, width=60).grid(row=1, column=1, sticky="w", padx=5)

        # 連番設定（互換）
        tk.Label(frm, text="章開始番号 (X)").grid(row=2, column=0, sticky="w")
        self.ch_start_var = tk.StringVar(value=str(self.cfg.chapter_start))
        tk.Entry(frm, textvariable=self.ch_start_var, width=10).grid(row=2, column=1, sticky="w", padx=5)

        tk.Label(frm, text="章数").grid(row=2, column=1, sticky="w", padx=110)
        self.ch_count_var = tk.StringVar(value=str(self.cfg.chapter_count))
        tk.Entry(frm, textvariable=self.ch_count_var, width=10).grid(row=2, column=1, sticky="w", padx=160)

        tk.Label(frm, text="合格: 各章スコア >= ").grid(row=3, column=0, sticky="w")
        self.pass_thr_var = tk.StringVar(value=str(self.cfg.pass_score_threshold))
        tk.Entry(frm, textvariable=self.pass_thr_var, width=10).grid(row=3, column=1, sticky="w", padx=5)

        tk.Label(frm, text="合格: 許可 >= ").grid(row=3, column=1, sticky="w", padx=110)
        self.perm_thr_var = tk.StringVar(value=str(self.cfg.permission_threshold))
        tk.Entry(frm, textvariable=self.perm_thr_var, width=10).grid(row=3, column=1, sticky="w", padx=210)

        tk.Label(frm, text="不合格: どれかの章スコア <= ").grid(row=4, column=0, sticky="w")
        self.fail_leq_var = tk.StringVar(value=str(self.cfg.fail_if_any_score_leq))
        tk.Entry(frm, textvariable=self.fail_leq_var, width=10).grid(row=4, column=1, sticky="w", padx=5)

        tk.Label(frm, text="判定列ヘッダー").grid(row=5, column=0, sticky="w")
        self.out_header_var = tk.StringVar(value=self.cfg.output_header)
        tk.Entry(frm, textvariable=self.out_header_var, width=20).grid(row=5, column=1, sticky="w", padx=5)

        tk.Label(frm, text="判定付きサフィックス").grid(row=5, column=1, sticky="w", padx=240)
        self.out_suffix_var = tk.StringVar(value=self.cfg.output_suffix)
        tk.Entry(frm, textvariable=self.out_suffix_var, width=20).grid(row=5, column=1, sticky="w", padx=360)

        tk.Label(frm, text="seiseki.xlsx（空なら target_dir/seiseki.xlsx を使用）").grid(row=6, column=0, sticky="w")
        self.seiseki_path_var = tk.StringVar(value=self.cfg.seiseki_path)
        tk.Entry(frm, textvariable=self.seiseki_path_var, width=95).grid(row=6, column=1, sticky="we", padx=5)
        tk.Button(frm, text="選択", command=self.choose_seiseki).grid(row=6, column=2, padx=5)

        tk.Label(frm, text="seiseki パスワード").grid(row=7, column=0, sticky="w")
        self.seiseki_pw_var = tk.StringVar(value=self.cfg.seiseki_password)
        tk.Entry(frm, textvariable=self.seiseki_pw_var, width=30, show="*").grid(row=7, column=1, sticky="w", padx=5)

        tk.Label(frm, text="対応キー: 判定付き(ELMS) / seiseki(学生番号)").grid(row=8, column=0, sticky="w")
        self.elms_header_var = tk.StringVar(value=self.cfg.judged_elms_header)
        self.stu_header_var = tk.StringVar(value=self.cfg.seiseki_student_header)
        tk.Entry(frm, textvariable=self.elms_header_var, width=20).grid(row=8, column=1, sticky="w", padx=5)
        tk.Entry(frm, textvariable=self.stu_header_var, width=20).grid(row=8, column=1, sticky="w", padx=190)

        tk.Label(frm, text="書込先ヘッダー（seiseki）").grid(row=9, column=0, sticky="w")
        self.eval_header_var = tk.StringVar(value=self.cfg.seiseki_eval_header)
        tk.Entry(frm, textvariable=self.eval_header_var, width=50).grid(row=9, column=1, sticky="w", padx=5)

        tk.Label(frm, text="seiseki出力サフィックス").grid(row=9, column=1, sticky="w", padx=520)
        self.seiseki_out_suffix_var = tk.StringVar(value=self.cfg.seiseki_output_suffix)
        tk.Entry(frm, textvariable=self.seiseki_out_suffix_var, width=20).grid(row=9, column=1, sticky="w", padx=680)

        tk.Button(frm, text="設定を読み込む", command=self.load_config).grid(row=10, column=0, pady=8, sticky="w")
        tk.Button(frm, text="設定を保存", command=self.save_config).grid(row=10, column=1, pady=8, sticky="w")

        tk.Button(frm, text="Step1: pXXX評定.xlsx を判定付きで保存", command=self.run_step1).grid(row=11, column=0, columnspan=3, pady=6, sticky="we")
        tk.Button(frm, text="Step2: 判定を seiseki.xlsx に反映して保存（未一致は警告して無視）", command=self.run_step2).grid(row=12, column=0, columnspan=3, pady=6, sticky="we")

        tk.Label(frm, text="ログ").grid(row=13, column=0, sticky="w")
        self.log = tk.Text(frm, height=20)
        self.log.grid(row=14, column=0, columnspan=3, sticky="nsew")

        frm.grid_columnconfigure(1, weight=1)
        frm.grid_rowconfigure(14, weight=1)

    def choose_dir(self):
        d = filedialog.askdirectory()
        if d:
            self.dir_var.set(d)

    def choose_seiseki(self):
        p = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if p:
            self.seiseki_path_var.set(p)

    def _sync_cfg_from_ui(self):
        self.cfg.target_dir = self.dir_var.get().strip()

        # 章番号リスト（優先）
        self.cfg.chapter_numbers = parse_chapter_numbers(self.ch_list_var.get().strip())

        # 互換：連番設定（章番号リストが空のときのみ使われる）
        self.cfg.chapter_start = int(self.ch_start_var.get().strip())
        self.cfg.chapter_count = int(self.ch_count_var.get().strip())

        self.cfg.pass_score_threshold = float(self.pass_thr_var.get().strip())
        self.cfg.permission_threshold = float(self.perm_thr_var.get().strip())
        self.cfg.fail_if_any_score_leq = float(self.fail_leq_var.get().strip())

        self.cfg.output_header = self.out_header_var.get().strip()
        self.cfg.output_suffix = self.out_suffix_var.get().strip()

        self.cfg.seiseki_path = self.seiseki_path_var.get().strip()
        self.cfg.seiseki_password = self.seiseki_pw_var.get()

        self.cfg.judged_elms_header = self.elms_header_var.get().strip()
        self.cfg.seiseki_student_header = self.stu_header_var.get().strip()
        self.cfg.seiseki_eval_header = self.eval_header_var.get().strip()
        self.cfg.seiseki_output_suffix = self.seiseki_out_suffix_var.get().strip()

        # 判定列ヘッダー（判定付き側）
        self.cfg.judged_result_header = self.cfg.output_header

        # ★安全策：ラベルは必ずG/F/5
        self.cfg.label_pass = "G"
        self.cfg.label_fail = "F"
        self.cfg.label_skip = "5"

    def _append_log(self, s: str):
        self.log.insert("end", s + "\n")
        self.log.see("end")

    def load_config(self):
        try:
            if not self.dir_var.get().strip():
                self.choose_dir()
                if not self.dir_var.get().strip():
                    return

            cfg_path = Path(self.dir_var.get().strip()) / self.cfg.config_filename
            if not cfg_path.exists():
                messagebox.showinfo("情報", f"設定ファイルが見つかりません:\n{cfg_path}")
                return

            with open(cfg_path, "r", encoding="utf-8") as f:
                data = json.load(f)

            self.cfg = safe_cfg_from_dict(data)

            self.dir_var.set(self.cfg.target_dir)
            self.ch_list_var.set(",".join(map(str, self.cfg.chapter_numbers)))
            self.ch_start_var.set(str(self.cfg.chapter_start))
            self.ch_count_var.set(str(self.cfg.chapter_count))

            self.pass_thr_var.set(str(self.cfg.pass_score_threshold))
            self.perm_thr_var.set(str(self.cfg.permission_threshold))
            self.fail_leq_var.set(str(self.cfg.fail_if_any_score_leq))

            self.out_header_var.set(self.cfg.output_header)
            self.out_suffix_var.set(self.cfg.output_suffix)

            self.seiseki_path_var.set(self.cfg.seiseki_path)
            self.seiseki_pw_var.set(self.cfg.seiseki_password)

            self.elms_header_var.set(self.cfg.judged_elms_header)
            self.stu_header_var.set(self.cfg.seiseki_student_header)
            self.eval_header_var.set(self.cfg.seiseki_eval_header)
            self.seiseki_out_suffix_var.set(self.cfg.seiseki_output_suffix)

            self._append_log(f"Loaded config: {cfg_path}")
        except Exception as e:
            messagebox.showerror("エラー", str(e))

    def save_config(self):
        try:
            self._sync_cfg_from_ui()
            cfg_path = Path(self.cfg.target_dir) / self.cfg.config_filename
            with open(cfg_path, "w", encoding="utf-8") as f:
                json.dump(asdict(self.cfg), f, ensure_ascii=False, indent=2)
            self._append_log(f"Saved config: {cfg_path}")
            messagebox.showinfo("完了", f"設定を保存しました:\n{cfg_path}")
        except Exception as e:
            messagebox.showerror("エラー", str(e))

    def run_step1(self):
        try:
            self._sync_cfg_from_ui()
            target = Path(self.cfg.target_dir)
            if not target.exists():
                raise FileNotFoundError("target_dir が存在しません")

            pfiles = find_pfiles(self.cfg)
            if not pfiles:
                self._append_log(f"対象ファイルなし (regex={self.cfg.filename_regex})")
                messagebox.showinfo("情報", "対象ファイルが見つかりませんでした")
                return

            self._append_log(f"Using chapters: {get_chapters(self.cfg)}")

            for f in pfiles:
                out = process_one_pfile(f, self.cfg)
                self._append_log(f"OK: {f.name} -> {out.name}")

            messagebox.showinfo("完了", f"Step1 完了: {len(pfiles)} 件")
        except Exception as e:
            messagebox.showerror("エラー", str(e))

    def run_step2(self):
        try:
            self._sync_cfg_from_ui()

            if self.cfg.seiseki_password == "":
                raise ValueError("seiseki のパスワードが空です")

            if self.cfg.seiseki_path:
                seiseki_path = Path(self.cfg.seiseki_path)
            else:
                seiseki_path = Path(self.cfg.target_dir) / "seiseki.xlsx"

            if not seiseki_path.exists():
                raise FileNotFoundError(f"seiseki.xlsx が見つかりません: {seiseki_path}")

            judged_files = find_judged_files(self.cfg)
            if not judged_files:
                raise FileNotFoundError("判定付きファイル（*_判定付き.xlsx）が見つかりません。Step1を先に実行してください。")

            merged: dict[str, str] = {}
            for jf in judged_files:
                id2res, logs = build_id_to_result_map(jf, self.cfg)
                for line in logs:
                    self._append_log(line)
                merged.update(id2res)

            self._append_log(f"Merged IDs total: {len(merged)}")

            out_path, logs2, missing_ids = apply_results_to_seiseki(
                seiseki_path, self.cfg.seiseki_password, merged, self.cfg
            )
            for line in logs2:
                self._append_log(line)

            # ★未一致は「でかく警告」するが続行は完了扱い
            if missing_ids:
                show = missing_ids[: self.cfg.max_log_missing]
                msg = (
                    f"警告：seiseki.xlsx 側に存在しない学生番号が {len(missing_ids)} 件ありました。\n\n"
                    f"これらは無視して処理を完了します。\n"
                    f"(先頭{self.cfg.max_log_missing}件まで表示)\n\n"
                    + "\n".join(show)
                )
                messagebox.showwarning("未一致学生番号あり", msg)

            messagebox.showinfo("完了", f"Step2 完了:\n{out_path}")
        except Exception as e:
            messagebox.showerror("エラー", str(e))


if __name__ == "__main__":
    App().mainloop()
